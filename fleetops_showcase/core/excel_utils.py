import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment


def export_invoices_excel(queryset, month_label):
    """Export DriverInvoice queryset to a formatted .xlsx file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Invoices {month_label}"

    # Header row with orange fill
    headers = [
        'NO.', 'Name', 'Phone', 'Cash', 'Main Orders', 'Addl. Orders', 'Hours', 'Work Date',
    ]
    orange_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = orange_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, invoice in enumerate(queryset, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=invoice.driver.full_name)
        ws.cell(row=row_idx, column=3, value=invoice.driver.phone)
        ws.cell(row=row_idx, column=4, value=float(invoice.cash))
        ws.cell(row=row_idx, column=5, value=invoice.main_orders)
        ws.cell(row=row_idx, column=6, value=invoice.additional_orders)
        ws.cell(row=row_idx, column=7, value=float(invoice.hours))
        ws.cell(row=row_idx, column=8, value=str(invoice.specified_date))

    # Auto-fit columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, max_length + 2)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="invoices_{month_label}.xlsx"'
    wb.save(response)
    return response


def export_archive_excel(queryset, label='archive'):
    """Export InvoiceArchive queryset to a formatted .xlsx file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Archive {label}"

    headers = [
        'NO.', 'Name', 'Phone', 'Cash', 'Main Orders', 'Addl. Orders', 'Hours', 'Work Date',
    ]
    orange_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = orange_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, archive in enumerate(queryset, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=archive.driver_name)
        ws.cell(row=row_idx, column=3, value=archive.driver.phone)
        ws.cell(row=row_idx, column=4, value=float(archive.cash))
        ws.cell(row=row_idx, column=5, value=archive.main_orders)
        ws.cell(row=row_idx, column=6, value=archive.additional_orders)
        ws.cell(row=row_idx, column=7, value=float(archive.hours))
        ws.cell(row=row_idx, column=8, value=str(archive.archive_date))

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, max_length + 2)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="archive_{label}.xlsx"'
    wb.save(response)
    return response

def export_talabat_excel(queryset, label='talabat'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Salary {label}"

    headers = [
        'Driver ID', 'Name', 'Total Orders', 'Total Salary KD', 'Deduction KD', 'Net Salary KD'
    ]
    orange_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = orange_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, record in enumerate(queryset, 2):
        ws.cell(row=row_idx, column=1, value=record.driver.employee_serial_number)
        ws.cell(row=row_idx, column=2, value=record.driver.full_name)
        ws.cell(row=row_idx, column=3, value=record.total_orders)
        ws.cell(row=row_idx, column=4, value=float(record.total_salary))
        ws.cell(row=row_idx, column=5, value=float(record.deduction))
        ws.cell(row=row_idx, column=6, value=float(record.net_salary))

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(15, max_length + 2)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="salary_{label}.xlsx"'
    wb.save(response)
    return response

def export_contract_excel(queryset, label='contract'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Salary {label}"

    headers = [
        'Name', 'Total Salary KD', 'Absent Days', 'Deduction KD', 'Net Salary KD', 'Remark'
    ]
    orange_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = orange_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, record in enumerate(queryset, 2):
        ws.cell(row=row_idx, column=1, value=record.name)
        ws.cell(row=row_idx, column=2, value=float(record.total_salary))
        ws.cell(row=row_idx, column=3, value=record.absent)
        ws.cell(row=row_idx, column=4, value=float(record.deduction))
        ws.cell(row=row_idx, column=5, value=float(record.net_salary))
        ws.cell(row=row_idx, column=6, value=record.remark)

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(15, max_length + 2)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="salary_{label}.xlsx"'
    wb.save(response)
    return response

def generate_excel_template(model_type):
    """Generate a template .xlsx file for bulk upload with existing data."""
    from .models import Driver, Profile, Deduction, DriverInvoice, TalabatSalaryDetail, ContractSalaryDetail
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    # Define headers and fetch data per model
    if model_type == 'driver':
        headers = [
            'Full Name', 'Email', 'Phone', 'Civil ID Number',
            'Passport Number', 'Vehicle Plate Number', 'Vehicle Name', 'Company Name',
            'Contract Type', 'Vehicle Type'
        ]
        queryset = Driver.objects.all()
        data_rows = []
        for obj in queryset:
            data_rows.append([
                obj.full_name, obj.email, obj.phone, obj.civil_id_number,
                obj.passport_number, obj.vehicle_plate_number, obj.vehicle_name,
                obj.company_name, obj.contract_type, obj.vehicle_type
            ])
    elif model_type == 'team':
        headers = [
            'First Name', 'Last Name', 'Email', 'Phone', 'Identification Number',
            'Passport', 'Role', 'Position', 'Base Salary KD'
        ]
        queryset = Profile.objects.all()
        data_rows = []
        for obj in queryset:
            data_rows.append([
                obj.first_name, obj.last_name, obj.email, obj.phone,
                obj.identification_number, obj.passport, obj.role, obj.position,
                float(obj.base_salary_kd)
            ])
    elif model_type == 'deduction':
        headers = [
            'Driver Email', 'Reason', 'Contracting Company', 
            'Contractor Deduction KD', 'Company Deduction KD'
        ]
        queryset = Deduction.objects.all()
        data_rows = []
        for obj in queryset:
            email = obj.driver.email if obj.driver else (obj.employee.email if obj.employee else "")
            data_rows.append([
                email, obj.reason, obj.contracting_company,
                float(obj.contractor_deduction_kd), float(obj.company_deduction_kd)
            ])
    elif model_type == 'invoice':
        headers = [
            'NO.', 'Name', 'Phone', 'Cash', 'Main Orders', 'Addl. Orders', 'Hours', 'Work Date'
        ]
        queryset = DriverInvoice.objects.all()
        data_rows = []
        for i, obj in enumerate(queryset, 1):
            data_rows.append([
                i, obj.driver.full_name, obj.driver.phone, float(obj.cash),
                obj.main_orders, obj.additional_orders, float(obj.hours),
                str(obj.specified_date)
            ])
    elif model_type == 'talabat_salary':
        headers = [
            'Driver ID', 'Batch 1 Orders', 'Batch 1 Amount', 'Batch 2 Orders', 'Batch 2 Amount',
            'Batch 3 Orders', 'Batch 3 Amount', 'Batch 4 Orders', 'Batch 4 Amount',
            'Batch 5 Orders', 'Batch 5 Amount', 'Batch 6 Orders', 'Batch 6 Amount',
            'Batch 7 Orders', 'Batch 7 Amount', 'Deduction'
        ]
        queryset = TalabatSalaryDetail.objects.all()
        data_rows = []
        for obj in queryset:
            data_rows.append([
                obj.driver.employee_serial_number, obj.batch_1_orders, float(obj.batch_1_amount),
                obj.batch_2_orders, float(obj.batch_2_amount), obj.batch_3_orders, float(obj.batch_3_amount),
                obj.batch_4_orders, float(obj.batch_4_amount), obj.batch_5_orders, float(obj.batch_5_amount),
                obj.batch_6_orders, float(obj.batch_6_amount), obj.batch_7_orders, float(obj.batch_7_amount),
                float(obj.deduction)
            ])
    elif model_type == 'contract_salary':
        headers = [
            'Name', 'Contract Type', 'Month', 'Total Salary', 'Absent Days', 'Deduction', 'Remark'
        ]
        queryset = ContractSalaryDetail.objects.all()
        data_rows = []
        for obj in queryset:
            data_rows.append([
                obj.name, obj.contract_type, str(obj.month), float(obj.total_salary),
                obj.absent, float(obj.deduction), obj.remark
            ])
    else:
        headers = ['Data']
        data_rows = []

    orange_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = orange_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 22

    # Populate Data Rows
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{model_type}_template.xlsx"'
    wb.save(response)
    return response


def import_from_excel(file, model_type, user):
    """Import data from uploaded Excel file."""
    from .models import Driver, Profile, DriverInvoice, Deduction
    from decimal import Decimal
    from datetime import date
    
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    rows = list(ws.rows)
    
    if len(rows) < 2:
        return 0, "File is empty or missing data."

    headers = [cell.value for cell in rows[0]]
    count = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], 2):
        data = {headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)}
        
        try:
            if model_type == 'driver':
                Driver.objects.create(
                    full_name=data.get('Full Name'),
                    email=data.get('Email'),
                    phone=data.get('Phone'),
                    civil_id_number=data.get('Civil ID Number'),
                    passport_number=data.get('Passport Number'),
                    vehicle_plate_number=data.get('Vehicle Plate Number'),
                    vehicle_name=data.get('Vehicle Name'),
                    company_name=data.get('Company Name', 'najmat'),
                    contract_type=data.get('Contract Type', 'talabat'),
                    vehicle_type=data.get('Vehicle Type', 'car'),
                    created_by=user.profile if hasattr(user, 'profile') else None
                )
            elif model_type == 'team':
                Profile.objects.create_user(
                    username=data.get('Email'),
                    email=data.get('Email'),
                    first_name=data.get('First Name'),
                    last_name=data.get('Last Name'),
                    phone=data.get('Phone', ''),
                    identification_number=data.get('Identification Number', ''),
                    passport=data.get('Passport', ''),
                    role=data.get('Role', 'employee'),
                    position=data.get('Position', 'Administrative'),
                    base_salary_kd=Decimal(str(data.get('Base Salary KD', 0) or 0)),
                    password='Password123!' # Default password
                )
            elif model_type == 'invoice':
                driver_phone = str(data.get('Phone', '')).strip()
                if not driver_phone:
                    errors.append(f"Row {row_idx}: Missing Phone number.")
                    continue
                
                try:
                    driver = Driver.objects.get(phone=driver_phone)
                except Driver.DoesNotExist:
                    errors.append(f"Row {row_idx}: Driver with phone {driver_phone} not found.")
                    continue
                
                DriverInvoice.objects.update_or_create(
                    driver=driver,
                    specified_date=data.get('Work Date') or date.today(),
                    defaults={
                        'cash': Decimal(str(data.get('Cash', 0) or 0)),
                        'main_orders': int(data.get('Main Orders', 0) or 0),
                        'additional_orders': int(data.get('Addl. Orders', 0) or 0),
                        'hours': Decimal(str(data.get('Hours', 0) or 0)),
                        'created_by': user.profile if hasattr(user, 'profile') else None
                    }
                )
            elif model_type == 'talabat_salary':
                from .models import TalabatSalaryDetail
                driver_id_val = str(data.get('Driver ID', '')).strip()
                if not driver_id_val:
                    errors.append(f"Row {row_idx}: Missing Driver ID.")
                    continue
                try:
                    driver = Driver.objects.get(employee_serial_number=driver_id_val)
                except Driver.DoesNotExist:
                    errors.append(f"Row {row_idx}: Driver with ID {driver_id_val} not found.")
                    continue
                
                TalabatSalaryDetail.objects.create(
                    driver=driver,
                    batch_1_orders=int(data.get('Batch 1 Orders', 0) or 0),
                    batch_1_amount=Decimal(str(data.get('Batch 1 Amount', 0) or 0)),
                    batch_2_orders=int(data.get('Batch 2 Orders', 0) or 0),
                    batch_2_amount=Decimal(str(data.get('Batch 2 Amount', 0) or 0)),
                    batch_3_orders=int(data.get('Batch 3 Orders', 0) or 0),
                    batch_3_amount=Decimal(str(data.get('Batch 3 Amount', 0) or 0)),
                    batch_4_orders=int(data.get('Batch 4 Orders', 0) or 0),
                    batch_4_amount=Decimal(str(data.get('Batch 4 Amount', 0) or 0)),
                    batch_5_orders=int(data.get('Batch 5 Orders', 0) or 0),
                    batch_5_amount=Decimal(str(data.get('Batch 5 Amount', 0) or 0)),
                    batch_6_orders=int(data.get('Batch 6 Orders', 0) or 0),
                    batch_6_amount=Decimal(str(data.get('Batch 6 Amount', 0) or 0)),
                    batch_7_orders=int(data.get('Batch 7 Orders', 0) or 0),
                    batch_7_amount=Decimal(str(data.get('Batch 7 Amount', 0) or 0)),
                    deduction=Decimal(str(data.get('Deduction', 0) or 0))
                )
            elif model_type == 'contract_salary':
                from .models import ContractSalaryDetail
                name = str(data.get('Name', '')).strip()
                if not name:
                    errors.append(f"Row {row_idx}: Missing Name.")
                    continue
                
                ContractSalaryDetail.objects.create(
                    name=name,
                    contract_type=data.get('Contract Type', 'other'),
                    month=data.get('Month', date.today().replace(day=1)),
                    total_salary=Decimal(str(data.get('Total Salary', 0) or 0)),
                    absent=int(data.get('Absent Days', 0) or 0),
                    deduction=Decimal(str(data.get('Deduction', 0) or 0)),
                    remark=str(data.get('Remark', ''))
                )
            count += 1
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")

    return count, errors
